import requests as req
import pandas as pd
import openpyxl
import base64
import json

xlsxFileName = "Pipelines.xlsx"
pipelinesSheetName = "Pipelines"

organizationName = ""
projectName = ""
PAT = ""

authorization = str(base64.b64encode(bytes(':' + PAT, 'ascii')), 'ascii')
headers = {
    'Accept': 'application/json',
    'Authorization': 'Basic ' + authorization
}

pipelinesUrl = f"https://dev.azure.com/{organizationName}/{projectName}/_apis/pipelines?api-version=7.0"

pipelinesResponse = req.get(url=pipelinesUrl, headers=headers)
pipelinesContent = json.loads(pipelinesResponse.text)

df = pd.DataFrame.from_dict(pipelinesContent["value"])

df.to_excel(xlsxFileName, pipelinesSheetName, index=False)

pipeline_file = openpyxl.load_workbook(xlsxFileName)
pipeline_list = pipeline_file[pipelinesSheetName]
repoColumnName = "Repository"
pipeline_list.cell(1, 7, value=repoColumnName)
pipeline_list.cell(1, 8, value="Trigger Branch")

for pipelines in range(2, pipeline_list.max_row + 1):

    pipelineId = pipeline_list.cell(pipelines, 3)
    pipeline_list.cell(pipelines, 1, value="")
    pipeline_list.cell(pipelines, 2, value="")

    pipelineUrl = f"https://dev.azure.com/{organizationName}/{projectName}/_apis/pipelines/{pipelineId.value}?api-version=7.0"
    detailedReq = req.get(url=pipelineUrl, headers=headers)
    pipelineReq = json.loads(detailedReq.text)
    configurationDict = pipelineReq["configuration"]
    for index in configurationDict:
        if index == "designerJson":
            designerDict = configurationDict["designerJson"]
            for repos in designerDict:
                if repos == "repository":
                    repoDict = designerDict["repository"]
                    for i in repoDict:
                        if i == "name":
                            pipeline_list.cell(pipelines, 7, value=str(repoDict[i]))
            for trigger in designerDict:
                if trigger == "triggers":
                    triggerList = designerDict["triggers"]
                    triggerDict = triggerList[0]
                    for i in triggerDict:
                        if i == "branchFilters":
                            pipeline_list.cell(pipelines, 8, value=str(triggerDict[i]))
        elif index == "path":
            pipeline_list.cell(pipelines, 7, value="See pipeline yaml file!")
            pipeline_list.cell(pipelines, 8, value="See pipeline yaml file!")


pipeline_file.save(xlsxFileName)
